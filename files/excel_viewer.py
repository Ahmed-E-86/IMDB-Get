import log
import subprocess
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import askyesno, showinfo, showerror
from tkinter.scrolledtext import ScrolledText
from tkinter import filedialog as fd
import sys
import os.path
import webbrowser
from PIL import Image, ImageTk

try:
    from openpyxl import load_workbook
    log.logger.debug('openpyxl is installed.')
except ImportError:
    log.logger.debug('openpyxl is not installed.')
    ask1 = askyesno(title='Import failed!', message='openpyxl is not installed. \nDo you want to install it?')
    if ask1:
        process1 = subprocess.run(['pip', 'install', 'openpyxl'])
    else:
        exit()
    if process1.returncode == 0:
        showinfo(title='Success!', message='openpyxl is installed successfully.')
        from openpyxl import load_workbook
    else:
        showerror(title='Error!', message='openpyxl installation is failed.')
        log.logger.error('openpyxl installation is failed.')
        exit()

# Get the directory of the current python file
main_dir = os.path.realpath(os.path.dirname(__file__))

# Setting the main window
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.geometry('1280x720')
        self.title("Excel viewer")
        # Change the theme
        self.tk.call("source", os.path.join(main_dir, "azure.tcl"))
        self.tk.call("set_theme", "light")
        # Set the icon
        icon = os.path.join(main_dir, 'images', 'logo.png')
        self.iconphoto(False, tk.PhotoImage(file=icon))
        # Grid configurations
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=2)
        self.columnconfigure(3, weight=2)

        self.Title, self.Year, self.Rating, self.AgeRating, self.Genre, self.Actors, self.Plot, self.Language, self.Awards, self.Poster, self.Url = ([] for list in range(11))
        self.dir = ''
        self.file = ''
        opened_file = sys.argv[1:]
        if opened_file == []:
            self.load_last_file()
        else:
            self.file = opened_file[0]
            self.dir = os.path.dirname(self.file)
            self.load_data(self.file)

        ## Label
        ttk.Label(self, text='Select a movie from the list below').grid(row=0, column=0, sticky='w', padx=5, pady=(5,0))

        ## Listbox
        self.listbox_var = tk.Variable(value=self.Title)
        self.listbox = tk.Listbox(self, listvariable=self.listbox_var, height=20, selectmode=tk.SINGLE, font=14)
        self.listbox.grid(row=1, column=0, sticky='nsew',rowspan=10, padx=5, pady=5, ipadx=5, ipady=5)
        self.listbox.bind('<<ListboxSelect>>', self.onselect)
        self.index = 0
        # Select the first item
        self.listbox.selection_set(self.index)

        ## Scrollbar
        scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.listbox.yview)
        self.listbox['yscrollcommand'] = scrollbar.set
        scrollbar.grid(row=1, column=0, sticky='ens', rowspan=10, padx=5, pady=5)

        ## Poster
        # Loading the image
        try:
            image = Image.open(self.Poster[self.index])
            img = ImageTk.PhotoImage(image)
        except FileNotFoundError:
            log.logger.error('The image file is not found.')
            img = None
        # Image label
        self.image_lbl = ttk.Label(self, borderwidth=5, text='Image not found', image=img)
        self.image_lbl.grid(row=1, column=3, columnspan=2, rowspan=10)
        self.image_lbl.image = img

        ## Title
        ttk.Label(self, text='Movie Name:', font=14).grid(row=1, column=1, sticky='w', padx=5, pady=5)
        self.title_lbl = ttk.Label(text=self.Title[self.index].title(), font=14, background='grey')
        self.title_lbl.grid(row=1, column=2, sticky='w', padx=5, pady=5)

        ## Year
        ttk.Label(self, text='Release Year:', font=14).grid(row=2, column=1, sticky='w', padx=5, pady=5)
        self.year_lbl = ttk.Label(text=self.Year[self.index], font=14, background='grey')
        self.year_lbl.grid(row=2, column=2, sticky='w', padx=5, pady=5)

        ## Rating
        ttk.Label(self, text='IMDB Rating:', font=14).grid(row=3, column=1, sticky='w', padx=5, pady=5)
        self.rating_lbl = ttk.Label(text=self.Rating[self.index], font=14, background='grey')
        self.rating_lbl.grid(row=3, column=2, sticky='w', padx=5, pady=5)

        ## Age Rating
        ttk.Label(self, text='Age Rating:', font=14).grid(row=4, column=1, sticky='w', padx=5, pady=5)
        self.age_lbl = ttk.Label(text=self.AgeRating[self.index], font=14, background='grey')
        self.age_lbl.grid(row=4, column=2, sticky='w', padx=5, pady=5)

        ## Genre
        ttk.Label(self, text='Genre:', font=14).grid(row=5, column=1, sticky='w', padx=5, pady=5)
        self.genre_lbl = ttk.Label(text=self.Genre[self.index], font=14, background='grey')
        self.genre_lbl.grid(row=5, column=2, sticky='w', padx=5, pady=5)

        ## Language
        ttk.Label(self, text='Language:', font=14).grid(row=6, column=1, sticky='w', padx=5, pady=5)
        self.lang_lbl = ttk.Label(text=self.Language[self.index], font=14, background='grey')
        self.lang_lbl.grid(row=6, column=2, sticky='w', padx=5, pady=5)

        ## Actors
        ttk.Label(self, text='Actors:', font=14).grid(row=7, column=1, sticky='w', padx=5, pady=5)
        self.actors_lbl = ttk.Label(text=self.Actors[self.index], font=14, wraplength=350, background='grey')
        self.actors_lbl.grid(row=7, column=2, sticky='w', padx=5, pady=5)

        ## Awards
        ttk.Label(self, text='Awards:', font=14).grid(row=8, column=1, sticky='w', padx=5, pady=5)
        self.awards_lbl = ttk.Label(text=self.Awards[self.index], font=14, background='grey')
        self.awards_lbl.grid(row=8, column=2, sticky='w', padx=5, pady=5)

        ## Plot
        ttk.Label(self, text='Plot:', font=14).grid(row=9, column=1, sticky='nw', padx=5, pady=5)
        self.plot_txt = ScrolledText(self, height=6, font=14, wrap='word', width=50)
        self.plot_txt.insert('end', self.Plot[self.index])
        self.plot_txt['state'] = 'disabled'
        self.plot_txt.grid(row=9, column=2, sticky='w', padx=5, pady=5)

        ## IMDB url
        ttk.Label(self, text='IMDB url:', font=14).grid(row=10, column=1, sticky='w', padx=5, pady=5)
        self.url_lbl = ttk.Label(text=self.Url[self.index], font=14, foreground='blue', cursor='hand2')
        self.url_lbl.grid(row=10, column=2, sticky='w', padx=5, pady=5)
        self.url_lbl.bind("<Button-1>", lambda e: self.callback(self.Url[self.index]))

        ## Next button
        img1 = ImageTk.PhotoImage(Image.open(os.path.join(main_dir, 'images/next.png')))
        self.btn1 = ttk.Button(self, text='Next', image=img1, compound='left', command=self.next)
        self.btn1.image = img1
        self.btn1.place(x=800, y=600)
        ## Previous button
        img2 = ImageTk.PhotoImage(Image.open(os.path.join(main_dir, 'images/previous.png')))
        self.btn2 = ttk.Button(self, text='Previous', image=img2, compound='right', command=self.previous)
        self.btn2.image = img2
        self.btn2.place(x=500, y=600)
        ## Load excel file button
        img3 = ImageTk.PhotoImage(Image.open(os.path.join(main_dir, 'images/open_xls.png')))
        self.btn3 = ttk.Button(self, text='Load excel file', image=img3, compound='right', command=self.load_excel_file)
        self.btn3.image = img3
        self.btn3.place(x=1100, y=600)
    
    def next(self):
        self.listbox.selection_clear(0, len(self.Title))
        if self.index == 0:
            self.index +=1
            self.listbox.selection_set(self.index)
            self.listbox.see(self.index)
            self.onselect(event=None)
        elif self.index < len(self.Title):
            self.index +=1
            self.listbox.selection_set(self.index)
            self.listbox.see(self.index)
            self.onselect(event=None)
        elif self.index == len(self.Title):
            self.index = 0
            self.listbox.selection_set(self.index)
            self.listbox.see(self.index)
            self.onselect(event=None)

    def previous(self):
        self.listbox.selection_clear(0, len(self.Title))
        if self.index == 0:
            self.index = len(self.Title)
            self.listbox.selection_set(self.index)
            self.listbox.see(self.index)
            self.onselect(event=None)
        elif self.index < len(self.Title):
            self.index -=1
            self.listbox.selection_set(self.index)
            self.listbox.see(self.index)
            self.onselect(event=None)
        elif self.index == len(self.Title):
            self.index -=1
            self.listbox.selection_set(self.index)
            self.listbox.see(self.index)
            self.onselect(event=None)
    
    def callback(self, url):
        webbrowser.open_new(url)
    
    def onselect(self, event):
        self.index = int(self.listbox.curselection()[0])
        self.update_data()
        
    def update_data(self):
        self.title_lbl['text'] = self.Title[self.index]
        self.year_lbl['text'] = self.Year[self.index]
        self.rating_lbl['text'] = self.Rating[self.index]
        self.age_lbl['text'] = self.AgeRating[self.index]
        self.genre_lbl['text'] = self.Genre[self.index]
        self.lang_lbl['text'] = self.Language[self.index]
        self.actors_lbl['text'] = self.Actors[self.index]
        self.awards_lbl['text'] = self.Awards[self.index]
        self.plot_txt['state'] = 'normal'
        self.plot_txt.delete(1.0, 'end')
        self.plot_txt.insert('end', self.Plot[self.index])
        self.plot_txt['state'] = 'disabled'
        self.url_lbl['text'] = self.Url[self.index]
        if self.Poster[self.index].endswith('A'):
            self.image_lbl['image'] = None
            self.image_lbl.image = None
        else:
            try:
                image = Image.open(self.Poster[self.index])
                img = ImageTk.PhotoImage(image)
                self.image_lbl['image'] = img
                self.image_lbl.image = img
            except FileNotFoundError:
                log.logger.error('The image file is not found.')
                self.image_lbl['image'] = None
                self.image_lbl.image = None
    
    def load_last_file(self):
        with open(os.path.join(main_dir, 'last_location'), 'r') as text_file:
            last_location = text_file.read()
        self.file = os.path.join(last_location, 'movies.xlsx')
        if os.path.isfile(self.file):
            self.dir = os.path.dirname(self.file)
            self.load_data(self.file)
        else:
            self.lower()
            showerror(title='File is not found!', message='The excel file is not found.')
            self.lift()
            self.file = self.open_file()
            self.dir = os.path.dirname(self.file)
            self.load_data(self.file)

    def load_data(self, excel_file):
        try:
            workbook = load_workbook(filename=excel_file)
        except FileNotFoundError:
            log.logger.error('The excel file is not found.')
            self.lower()
            showerror(title='Error', message='The excel file is not found, \nthe application is exiting now.')
            exit()
        # Load the first sheet in the workbook
        sheet = workbook.active
        # Get all the data inside the sheet row by row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row != []:
                title = row[1].capitalize()
                self.Title.append(title)
                year = row[2]
                self.Year.append(year)
                rating = row[3]
                self.Rating.append(rating)
                age_rating = row[4]
                self.AgeRating.append(age_rating)
                genre = row[5]
                self.Genre.append(genre)
                actors = row[6]
                self.Actors.append(actors)
                plot = row[7]
                self.Plot.append(plot)
                language = row[8]
                self.Language.append(language)
                awards = row[9]
                self.Awards.append(awards)
                poster = row[10]
                # Extract the filename from the url
                filename = poster.split('/')[-1]
                # Add the path to the filename
                file_path = os.path.join(self.dir, "(posters)", filename)
                self.Poster.append(file_path)
                imdb_url = row[11]
                self.Url.append(imdb_url)

    def open_file(self):
        filetypes = (('xlsx files', 'movies.xlsx'), ('All files', '*.*'))
        self.lower()
        open_file = fd.askopenfile(title='Open excel file', initialdir='/', filetypes=filetypes)
        self.lift()
        # Check if the open dialog is canceled
        try:
            excel_file = open_file.name
        except AttributeError:
            log.logger.debug('Open dialog is cancelled, try again.')
            self.open_file()
        # Get the location from the excel file to be used later to load the posters
        self.dir = os.path.dirname(excel_file)
        return excel_file
    
    def load_excel_file(self):
        excel_file = self.open_file()
        print(excel_file)
        # Clear all the lists
        self.Title, self.Year, self.Rating, self.AgeRating, self.Genre, self.Actors, self.Plot, self.Language, self.Awards, self.Poster, self.Url = ([] for list in range(11))
        # Open the excel file and update the lists
        self.load_data(excel_file)
        # Update the fields
        self.update_data()
        # Clear the listbox
        self.listbox.delete(0, 'end')
        # Update the listbox
        for item in self.Title:
            self.listbox.insert('end', item)
        # Change the current selection
        self.index = 0
        self.listbox.selection_set(self.index)
        # Overwrie the last location
        with open(os.path.join(main_dir, 'last_location'), 'w') as text_file:
            text_file.write(self.dir)

if __name__ == "__main__":
    exc_viewer = App()
    exc_viewer.mainloop()